#!/usr/bin/env python3
"""Validate and upsert Singapore semiconductor company profiles.

The AI agent researches the website and returns one JSON object. This script
owns the deterministic parts: domain normalization, duplicate checks, Pydantic
validation, and Excel writes.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

SKILL_DIR = Path(__file__).resolve().parents[1]
if str(SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(SKILL_DIR))

from schemas.company_profile import COMPANY_COLUMNS, CompanyProfile  # noqa: E402
from schemas.company_profile import normalize_domain as _normalize_domain  # noqa: E402


WORKBOOK_ENV = "COMPANIES_XLSX"
PROJECT_ROOT_ENV = "SEMICON_PROJECT_ROOT"
DEFAULT_WORKBOOK = Path("data/companies.xlsx")
SHEET_NAME = "companies"
PRIMARY_KEY = "domain"
STALE_AFTER_DAYS = 90


def project_root() -> Path:
    return Path(os.environ.get(PROJECT_ROOT_ENV, Path.cwd())).resolve()


def workbook_path() -> Path:
    root = project_root()
    configured = Path(os.environ.get(WORKBOOK_ENV, DEFAULT_WORKBOOK))
    path = configured if configured.is_absolute() else root / configured
    resolved = path.resolve()

    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"workbook path must stay inside project root: {resolved}") from exc

    return resolved


def normalize_domain(url: str) -> str:
    return _normalize_domain(url)


def _load_or_create_workbook() -> tuple[Workbook, Worksheet]:
    path = workbook_path()
    if path.exists():
        workbook = load_workbook(path)
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()

    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.active
        sheet.title = SHEET_NAME

    _ensure_headers(sheet)
    return workbook, sheet


def _load_existing_workbook() -> tuple[Workbook, Worksheet] | None:
    path = workbook_path()
    if not path.exists():
        return None
    workbook = load_workbook(path)
    if SHEET_NAME not in workbook.sheetnames:
        return None
    sheet = workbook[SHEET_NAME]
    changed = _ensure_headers(sheet)
    if changed:
        _save_workbook(workbook)
    return workbook, sheet


def _save_workbook(workbook: Workbook) -> None:
    path = workbook_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=path.parent,
            prefix=f".{path.stem}-",
            suffix=path.suffix,
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink()


def _ensure_headers(sheet: Worksheet) -> bool:
    headers = [cell.value for cell in sheet[1]]
    if not any(headers):
        for index, column in enumerate(COMPANY_COLUMNS, start=1):
            sheet.cell(row=1, column=index, value=column)
        sheet.freeze_panes = "A2"
        last_column = sheet.cell(row=1, column=len(COMPANY_COLUMNS)).column_letter
        sheet.auto_filter.ref = f"A1:{last_column}1"
        return True

    missing = [column for column in COMPANY_COLUMNS if column not in headers]
    if missing:
        next_column = len(headers) + 1
        for offset, column in enumerate(missing):
            sheet.cell(row=1, column=next_column + offset, value=column)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return bool(missing)


def _header_index(sheet: Worksheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1] if cell.value}


def workbook_audit() -> dict[str, Any]:
    path = workbook_path()
    result: dict[str, Any] = {
        "workbook": str(path),
        "exists": path.exists(),
        "sheet": SHEET_NAME,
        "required_columns": COMPANY_COLUMNS,
        "missing_columns": COMPANY_COLUMNS,
        "extra_columns": [],
        "row_count": 0,
        "duplicate_domains": [],
        "research_quality_counts": {},
        "rows_missing_evidence_urls": [],
    }
    loaded = _load_existing_workbook()
    if loaded is None:
        return result

    _, sheet = loaded
    headers = [cell.value for cell in sheet[1] if cell.value]
    result["missing_columns"] = [column for column in COMPANY_COLUMNS if column not in headers]
    result["extra_columns"] = [column for column in headers if column not in COMPANY_COLUMNS]
    result["row_count"] = max(sheet.max_row - 1, 0)

    if PRIMARY_KEY in headers:
        columns = _header_index(sheet)
        domain_col = columns[PRIMARY_KEY]
        counts: dict[str, int] = {}
        quality_counts: dict[str, int] = {}
        for row_num in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_num, column=domain_col).value
            if not value:
                continue
            domain = str(value).strip().lower()
            counts[domain] = counts.get(domain, 0) + 1

            if "research_quality" in columns:
                quality = sheet.cell(row=row_num, column=columns["research_quality"]).value
                quality_key = str(quality).strip() if quality else "missing"
                quality_counts[quality_key] = quality_counts.get(quality_key, 0) + 1

            if "evidence_urls" in columns:
                evidence_urls = sheet.cell(row=row_num, column=columns["evidence_urls"]).value
                if not evidence_urls or not str(evidence_urls).strip():
                    result["rows_missing_evidence_urls"].append(domain)

        result["duplicate_domains"] = [
            domain for domain, count in sorted(counts.items()) if count > 1
        ]
        result["research_quality_counts"] = dict(sorted(quality_counts.items()))

    return result


def _find_row(sheet: Worksheet, domain: str) -> int | None:
    rows = _find_rows(sheet, domain)
    return rows[0] if rows else None


def _find_rows(sheet: Worksheet, domain: str) -> list[int]:
    columns = _header_index(sheet)
    domain_col = columns[PRIMARY_KEY]
    target = domain.strip().lower()
    rows: list[int] = []
    for row_num in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_num, column=domain_col).value
        if value and str(value).strip().lower() == target:
            rows.append(row_num)
    return rows


def _parse_checked_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _row_is_stale(sheet: Worksheet, row_num: int, today: date | None = None) -> bool:
    columns = _header_index(sheet)
    checked = _parse_checked_date(sheet.cell(row=row_num, column=columns["last_checked"]).value)
    if checked is None:
        return True
    reference_date = today or date.today()
    return (reference_date - checked).days > STALE_AFTER_DAYS


def _row_missing_evidence_fields(sheet: Worksheet, row_num: int) -> bool:
    columns = _header_index(sheet)
    for column in ("evidence_url", "evidence_urls", "evidence_summary", "research_quality"):
        if column not in columns:
            return True
        value = sheet.cell(row=row_num, column=columns[column]).value
        if not value or not str(value).strip():
            return True
    return False


def _delete_company_row(sheet: Worksheet, domain: str) -> bool:
    row_num = _find_row(sheet, domain)
    if row_num is None:
        return False
    sheet.delete_rows(row_num, 1)
    sheet.auto_filter.ref = sheet.dimensions
    return True


def delete_company(domain: str) -> bool:
    loaded = _load_existing_workbook()
    if loaded is None:
        return False

    workbook, sheet = loaded
    deleted = _delete_company_row(sheet, domain)
    if deleted:
        _save_workbook(workbook)
    return deleted


def company_exists(domain: str) -> bool:
    loaded = _load_existing_workbook()
    if loaded is None:
        return False
    _, sheet = loaded
    return _find_row(sheet, domain) is not None


def read_company(domain: str) -> dict[str, Any] | None:
    loaded = _load_existing_workbook()
    if loaded is None:
        return None

    _, sheet = loaded
    row_num = _find_row(sheet, domain)
    if row_num is None:
        return None

    columns = _header_index(sheet)
    return {
        column: sheet.cell(row=row_num, column=columns[column]).value
        for column in COMPANY_COLUMNS
    }


def validate_company(data: dict[str, Any]) -> CompanyProfile:
    payload = dict(data)
    payload["domain"] = normalize_domain(payload.get("domain") or payload.get("website", ""))
    if "website" in payload:
        payload["website"] = str(payload["website"])
    if "evidence_url" in payload:
        payload["evidence_url"] = str(payload["evidence_url"])
    if "evidence_urls" in payload:
        payload["evidence_urls"] = _normalize_evidence_urls(payload["evidence_urls"])
    return CompanyProfile.model_validate(payload)


def _normalize_evidence_urls(value: Any) -> list[str]:
    if isinstance(value, str):
        urls = [
            item.strip()
            for chunk in value.splitlines()
            for item in chunk.split(";")
            if item.strip()
        ]
    elif isinstance(value, list):
        urls = [str(item).strip() for item in value if str(item).strip()]
    else:
        raise ValueError("evidence_urls must be a list of URLs or a newline-separated string")
    return list(dict.fromkeys(urls))


def _excel_value(value: Any) -> Any:
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return value


def upsert_company(profile: CompanyProfile) -> None:
    workbook, sheet = _load_or_create_workbook()
    columns = _header_index(sheet)

    row_nums = _find_rows(sheet, profile.domain)
    row_num = row_nums[0] if row_nums else None
    if row_num is not None and _row_is_stale(sheet, row_num):
        rows_to_delete = row_nums
        row_num = None
    else:
        rows_to_delete = row_nums[1:]

    for duplicate_row in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(duplicate_row, 1)

    if row_num is None:
        row_num = sheet.max_row + 1

    row_data = profile.model_dump(mode="json")
    for column in COMPANY_COLUMNS:
        sheet.cell(row=row_num, column=columns[column], value=_excel_value(row_data.get(column, "")))

    sheet.auto_filter.ref = sheet.dimensions
    _save_workbook(workbook)


def check_url(url: str) -> str:
    domain = normalize_domain(url)
    loaded = _load_existing_workbook()
    if loaded is None:
        return f"research required: {domain}"

    _, sheet = loaded
    row_num = _find_row(sheet, domain)
    if row_num is None:
        return f"research required: {domain}"
    if _row_missing_evidence_fields(sheet, row_num):
        return f"research required: {domain} (missing evidence fields)"
    if _row_is_stale(sheet, row_num):
        return f"research required: {domain} (stale row retained until replacement)"
    if row_num:
        return "already exists"
    return f"research required: {domain}"


def process_url(url: str) -> str:
    return check_url(url)


def _load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise ValueError("profile JSON must be one object")
    return data


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Check, validate, and upsert CompanyProfile rows in data/companies.xlsx."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    check_parser = subparsers.add_parser("check", help="Check whether a URL needs AI research")
    check_parser.add_argument("url")

    read_parser = subparsers.add_parser("read", help="Read one existing row by domain or URL")
    read_parser.add_argument("url_or_domain")

    subparsers.add_parser("audit", help="Report workbook columns, row count, and duplicate domains")

    validate_parser = subparsers.add_parser("validate", help="Validate one profile JSON file")
    validate_parser.add_argument("json_path", type=Path)

    upsert_parser = subparsers.add_parser("upsert", help="Validate and upsert one profile JSON file")
    upsert_parser.add_argument("json_path", type=Path)

    args = parser.parse_args()

    try:
        if args.command == "check":
            print(check_url(args.url))
            return 0

        if args.command == "read":
            domain = normalize_domain(args.url_or_domain)
            row = read_company(domain)
            print(json.dumps(row, indent=2, ensure_ascii=False, default=str))
            return 0 if row else 1

        if args.command == "audit":
            print(json.dumps(workbook_audit(), indent=2, ensure_ascii=False, default=str))
            return 0

        if args.command == "validate":
            profile = validate_company(_load_json(args.json_path))
            print(profile.model_dump_json(indent=2))
            return 0

        if args.command == "upsert":
            profile = validate_company(_load_json(args.json_path))
            upsert_company(profile)
            print(f"upserted: {profile.domain}")
            return 0

    except (OSError, ValueError, ValidationError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    return 1


if __name__ == "__main__":
    sys.exit(main())
