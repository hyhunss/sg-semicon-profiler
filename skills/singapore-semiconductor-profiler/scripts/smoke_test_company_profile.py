#!/usr/bin/env python3
"""Smoke test for the Singapore Semiconductor Profiler skill."""

from __future__ import annotations

import importlib.util
import os
import tempfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


SKILL_DIR = Path(__file__).resolve().parents[1]
EXCEL_SCRIPT = SKILL_DIR / "scripts" / "company_profile_excel.py"


def load_excel_module():
    spec = importlib.util.spec_from_file_location("company_profile_excel", EXCEL_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def main() -> int:
    excel = load_excel_module()

    profiles = [
        {
            "company_name": "OneSystems Technologies Pte Ltd",
            "website": "https://www.onesystemstech.com",
            "domain": "onesystemstech.com",
            "business_summary": "OneSystems Technologies provides turnkey OT and IT infrastructure, cleanroom, communication, security, and integration solutions for semiconductor plants.",
            "semicon_role": "systems_integrator",
            "products_services": "OT/IT integration; cleanroom infrastructure; communications; security systems",
            "target_customer_type": "Semiconductor fabs; advanced manufacturing facilities; cleanroom operators",
            "buyer_need": "OT/IT integration; cleanroom infrastructure; facility communications; security systems",
            "evidence_url": "https://www.onesystemstech.com/semiconductor",
            "evidence_urls": [
                "https://www.onesystemstech.com/semiconductor",
                "https://www.onesystemstech.com/turnkey-infrastructure",
                "https://www.onesystemstech.com/about-us",
            ],
            "evidence_summary": "Semiconductor page supports plant infrastructure relevance; turnkey infrastructure page supports OT/IT and cleanroom services; about page supports company context.",
            "confidence": "high",
            "research_quality": "complete",
            "last_checked": "2026-05-26",
            "notes": "Direct semiconductor facility infrastructure evidence.",
        },
        {
            "company_name": "Intel Singapore",
            "website": "https://www.intel.com",
            "domain": "intel.com",
            "business_summary": "Intel designs and manufactures semiconductor products and is the canonical integrated device manufacturer example.",
            "semicon_role": "idm",
            "products_services": "Semiconductor design; wafer fabrication; processors; semiconductor manufacturing technology",
            "target_customer_type": "Computer manufacturers; cloud providers; electronics OEMs",
            "buyer_need": "Processors; semiconductor components; advanced compute platforms",
            "evidence_url": "https://www.intel.com",
            "evidence_urls": [
                "https://www.intel.com",
                "https://www.intel.com/content/www/us/en/products/details/processors.html",
                "https://www.intel.com/content/www/us/en/company-overview/company-overview.html",
            ],
            "evidence_summary": "Homepage and company overview support Intel identity; processor page supports semiconductor product offerings and customer needs.",
            "confidence": "high",
            "research_quality": "complete",
            "last_checked": "2026-05-26",
            "notes": "Smoke-test row for the IDM role.",
        },
    ]

    with tempfile.TemporaryDirectory() as tmp:
        os.environ.pop(excel.WORKBOOK_ENV, None)
        os.environ[excel.PROJECT_ROOT_ENV] = tmp
        expected_path = (Path(tmp) / "data" / "companies.xlsx").resolve()
        assert excel.workbook_path() == expected_path

        expected_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = excel.SHEET_NAME
        legacy_headers = [
            "company_name",
            "website",
            "domain",
            "country",
            "business_summary",
            "semicon_category",
            "products_services",
            "evidence_url",
            "evidence_urls",
            "evidence_summary",
            "confidence",
            "research_quality",
            "last_checked",
            "status",
            "notes",
        ]
        for index, header in enumerate(legacy_headers, start=1):
            sheet.cell(row=1, column=index, value=header)
        workbook.save(expected_path)

        for data in profiles:
            excel.upsert_company(excel.validate_company(data))

        workbook = load_workbook(expected_path)
        sheet = workbook[excel.SHEET_NAME]
        headers = [cell.value for cell in sheet[1] if cell.value]
        assert not [column for column in excel.COMPANY_COLUMNS if column not in headers]
        assert "country" in headers
        assert "status" in headers

        assert excel.company_exists("onesystemstech.com")
        assert excel.company_exists("intel.com")
        assert excel.check_url("https://www.onesystemstech.com") == "already exists"
        assert excel.read_company("onesystemstech.com")["semicon_role"] == "systems_integrator"
        audit = excel.workbook_audit()
        assert audit["row_count"] == 2
        assert audit["duplicate_domains"] == []
        assert audit["missing_columns"] == []
        assert audit["research_quality_counts"] == {"complete": 2}
        assert audit["rows_missing_evidence_urls"] == []
        assert "country" in audit["extra_columns"]
        assert "status" in audit["extra_columns"]

        stale_profile = profiles[0] | {
            "last_checked": (date.today() - timedelta(days=91)).isoformat(),
            "notes": "Stale row for replacement test.",
        }
        excel.upsert_company(excel.validate_company(stale_profile))
        assert excel.check_url("https://www.onesystemstech.com") == (
            "research required: onesystemstech.com (stale row retained until replacement)"
        )
        assert excel.company_exists("onesystemstech.com")

        fresh_profile = profiles[0] | {"last_checked": date.today().isoformat()}
        excel.upsert_company(excel.validate_company(fresh_profile))
        assert excel.check_url("https://www.onesystemstech.com") == "already exists"

        workbook = load_workbook(expected_path)
        sheet = workbook[excel.SHEET_NAME]
        duplicate_row = sheet.max_row + 1
        columns = excel._header_index(sheet)
        source_row = excel._find_row(sheet, "onesystemstech.com")
        assert source_row is not None
        for column in excel.COMPANY_COLUMNS:
            sheet.cell(
                row=duplicate_row,
                column=columns[column],
                value=sheet.cell(row=source_row, column=columns[column]).value,
            )
        workbook.save(expected_path)
        assert excel.workbook_audit()["duplicate_domains"] == ["onesystemstech.com"]

        excel.upsert_company(excel.validate_company(profiles[0] | {"notes": "Duplicate collapse."}))
        workbook = load_workbook(expected_path)
        sheet = workbook[excel.SHEET_NAME]
        columns = excel._header_index(sheet)
        domains = [
            sheet.cell(row=row_num, column=columns[excel.PRIMARY_KEY]).value
            for row_num in range(2, sheet.max_row + 1)
        ]
        assert domains.count("onesystemstech.com") == 1

        wrong_domain = profiles[0] | {"domain": "wrong.com.sg"}
        try:
            excel.validate_company(wrong_domain)
        except ValueError:
            pass
        else:
            raise AssertionError("mismatched website/domain should fail validation")

        weak_high_confidence = profiles[0] | {
            "evidence_urls": ["https://www.onesystemstech.com/semiconductor"],
            "research_quality": "thin_evidence",
        }
        try:
            excel.validate_company(weak_high_confidence)
        except ValueError:
            pass
        else:
            raise AssertionError("high confidence should require complete multi-page evidence")

        limited_profile = profiles[0] | {
            "confidence": "medium",
            "research_quality": "limited_site",
            "evidence_urls": "https://www.onesystemstech.com/semiconductor",
            "evidence_summary": "Only one relevant semiconductor page was available; role is supported but site coverage is limited.",
            "notes": "Limited evidence accepted at medium confidence.",
        }
        assert excel.validate_company(limited_profile).research_quality == "limited_site"

        external_high_confidence = profiles[0] | {
            "evidence_urls": [
                "https://www.onesystemstech.com/semiconductor",
                "https://example.com/source-a",
                "https://example.org/source-b",
            ],
        }
        try:
            excel.validate_company(external_high_confidence)
        except ValueError:
            pass
        else:
            raise AssertionError("complete high-confidence evidence should require company-site URLs")

        workbook = load_workbook(expected_path)
        sheet = workbook[excel.SHEET_NAME]
        columns = excel._header_index(sheet)
        legacy_row = excel._find_row(sheet, "onesystemstech.com")
        assert legacy_row is not None
        for column in ("evidence_urls", "evidence_summary", "research_quality"):
            sheet.cell(row=legacy_row, column=columns[column], value="")
        workbook.save(expected_path)
        assert excel.check_url("https://www.onesystemstech.com") == (
            "research required: onesystemstech.com (missing evidence fields)"
        )

        extra_fields = profiles[0] | {"country": "Singapore", "status": "enriched"}
        try:
            excel.validate_company(extra_fields)
        except ValueError:
            pass
        else:
            raise AssertionError("extra schema fields should fail validation")

        os.environ[excel.WORKBOOK_ENV] = "../outside.xlsx"
        try:
            excel.workbook_path()
        except ValueError:
            pass
        else:
            raise AssertionError("workbook path outside project root should fail")

    print("OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
